#!/usr/bin/env python
# encoding: utf-8
"""
    pyQms
    -----

    Python module for fast and accurate mass spectrometry data quantification

    :license: MIT, see LICENSE.txt for more details

    Authors:

        * Leufken, J.
        * Niehues, A.
        * Sarin, L.P.
        * Hippler, M.
        * Leidel, S.A.
        * Fufezan, C.

"""
from __future__ import absolute_import
from collections import namedtuple
import math
import re
import os
from operator import itemgetter
import pyqms
import bisect
import sys
import csv
import codecs
import pyqms.adaptors
import pprint
import copy
from collections import defaultdict as ddict
import pandas as pd

# import numpy

m_key = namedtuple("m_key", ["file_name", "formula", "charge", "label_percentiles"])

match = namedtuple("match", ["spec_id", "rt", "score", "scaling_factor", "peaks"])

combined = namedtuple(
    "combined",
    [
        "file_name",
        "formula",
        "charge",
        "label_percentiles",
        "spec_id",
        "rt",
        "score",
        "scaling_factor",
        "peaks",
    ],
)

"""
file_name, molecule, charge, label_percentile_tuple = key
"""

default_amount_csv_fieldnames = [
    "file_name",
    "formula",
    "molecule",
    "trivial_name(s)",
    "label_percentiles",
    "charge",
    "start (min)",
    "stop (min)",
    "max I in window",
    "max I in window (rt)",
    "max I in window (score)",
    "auc in window",
    "sum I in window",
    "evidences (min)",
]


class Results(dict):
    """
    pyQms results class.

    Holds all matching information and lookups. Can be accessed as a
    dictionary. Several lookup allow the mapping of moleclar formulas to
    molecules (e.g. peptides) and/or trivial names (e.g. protein names).


    Structure

        key (named tuple)

            * file_name
            * formula
            * charge
            * label_percentiles

        value (named tuple)

            * spec_id
            * rt
            * score
            * scaling_factor
            * peaks

    """

    def __init__(
        self,
        lookup=None,
        params=None,
        fixed_labels=None,
        metabolic_labels=None,
        aa_compositions=None,
        isotopic_distributions=None,
        charges=None,
        verbose=False,
    ):

        self.params = params
        if lookup is not None:
            self.lookup = lookup
        else:
            self.lookup = {}
        self.fixed_labels = fixed_labels
        self.metabolic_labels = metabolic_labels
        self.aa_compositions = aa_compositions
        self.charges = charges
        self.isotopic_distribution = isotopic_distributions
        self.verbose = verbose
        self.index = {
            "files": set(),
            "charges": set(),
            "formulas": set(),
            "label_percentiles": set(),
        }
        self._match_class = match
        self._m_key_class = m_key
        self._combined_class = combined
        self._silac_pairs = None
        return

    def add(self, key, value):
        """Adds match to the result container.

        Args:
            key (named tuple): ( file_name, formula, charge, label_percentiles )
            value (named tuple): (spec_id, rt, score, scaling_factor, peaks)

        Returns
            formatted_key (named tuple) : ( file_name, formula, charge, label_percentiles )

        Structure

            key (named tuple)

                * file_name
                * formula
                * charge
                * label_percentiles

            value (named tuple)

                * spec_id
                * rt
                * score
                * scaling_factor
                * peaks

        """
        m_key = self._m_key_class(*key)
        try:
            self[m_key]
        except:
            self[m_key] = {
                "data": [],
                "max_score": -1,
                "max_score_index": -1,
                "len_data": 0,
            }
        entry = self._match_class(*value)
        if self[m_key]["max_score"] < entry.score:
            self[m_key]["max_score"] = entry.score
            self[m_key]["max_score_index"] = len(self[m_key]["data"])
            # test for this ...
        self[m_key]["len_data"] += 1
        self[m_key]["data"].append(entry)
        self.index["files"].add(m_key.file_name)
        self.index["charges"].add(m_key.charge)
        self.index["formulas"].add(m_key.formula)
        self.index["label_percentiles"].add(m_key.label_percentiles)
        if self.params is not None and self.params["BUILD_RESULT_INDEX"] is True:
            # molecule -> m_key list
            # file     -> m_key list (index list)
            # try:
            #     self.index[ m_key.formula ]
            # except:
            #     self.index[ m_key.formula ] = set()

            # this should make max results obsolete
            pass
        return m_key

    def _parse_and_filter(
        self,
        molecules=None,
        charges=None,
        file_names=None,
        label_percentiles=None,
        formulas=None,
    ):
        """Filter keys in results which meet certain criteria.

        Generalized generator that filters the results and yields only those
        keys that match given criteria. If a parameter is equal `None` then
        it is not used to filter the data.

        Args:
            molecules (list of str, optional): considered molecules. Those will
                be translated using self._translate_molecules_to_formulas()
            charges (list of int, optional): considered charge
                states.
            file_names (list of str, optional): list of file names to be
                considered.
            label_percentiles (list of tuple, optional): list of label percentile tuples
                to be considered.
            formulas (list of str): list of chemical formulas

        Yields:
            key (tuple): that matches filters. Key is a named_tuple and has the
                following attributes:
                file_name, molecule, charge, label_percentile

        """
        if molecules is not None:
            formulas = self._translate_molecules_to_formulas(molecules, formulas)
        for key in self.keys():
            # file_name, molecule, charge, label_percentile_tuple = key
            if file_names is not None and key[0] not in file_names:
                continue
            if formulas is not None and key[1] not in formulas:
                continue
            if charges is not None and key[2] not in charges:
                continue
            if label_percentiles is not None and key[3] not in label_percentiles:
                continue
            yield key

    def extract_results(
        self,
        molecules=None,
        charges=None,
        file_names=None,
        label_percentiles=None,
        formulas=None,
        score_threshold=None,
    ):
        """
        Extract selected results.

        Extracts all matches from the results instance that meet given filter
        criteria.

        Args:
            molecules (list of str, optional): considered molecules. Those will
                be translated using self._translate_molecules_to_formulas()
            charges (list of int, optional): considered charge
                states.
            file_names (list of str, optional): list of file names to be
                considered.
            label_percentiles (list of tuple, optional): list of label percentile tuples
                to be considered.
            formulas (list of str): list of chemical formulas

        Yields:
            key, i, entry (tuple) : result class key, index of entry and entry

        Structure

            key (named tuple)

                * file_name
                * formula
                * charge
                * label_percentiles

            value (named tuple)

                * spec_id
                * rt
                * score
                * scaling_factor
                * peaks

        """
        for key in self._parse_and_filter(
            molecules=molecules,
            charges=charges,
            file_names=file_names,
            label_percentiles=label_percentiles,
            formulas=formulas,
        ):
            for i, entry in enumerate(self[key]["data"]):
                if score_threshold is not None and entry.score < score_threshold:
                    continue
                yield key, i, entry

    def format_all_results(
        self,
    ):
        """
        Format all results to a pandas DataFrame.

        Args:

        returns:
            results_df (Dataframe) : containing all results with their m_key data

        Structure

            columns
                * file_name
                * formula
                * charge
                * label_percentiles
                * spec_id
                * rt
                * score
                * scaling_factor
                * peaks

        """
        results = []
        for key in self.keys():
            for ion_match in self[key]["data"]:
                extended = self._combined_class(*key, *ion_match)
                results.append(extended)

        results_df = pd.DataFrame(results)
        return results_df

    def _translate_molecules_to_formulas(self, molecules, formulas):
        """"""
        if formulas is not None:
            formulas = set(formulas)
        else:
            formulas = set()
        for molecule in molecules:
            try:
                translation = self.lookup["molecule to formula"][molecule]
                formulas.add(translation)
            except:
                pass
                # print('Cannot find formula translation for {0}'.format(
                #     molecule )
                # )
        return formulas

    def max_score(
        self,
        molecules=None,
        charges=None,
        file_names=None,
        label_percentiles=None,
        formulas=None,
    ):
        """Find max score for a given set of parameters.

        Args:
            molecules (list of str, optional): considered molecules. Those will
                be translated using self._translate_molecules_to_formulas()
            charges (list of int, optional): considered charge
                states.
            file_names (list of str, optional): list of file names to be
                considered.
            label_percentiles (list of tuple, optional): list of label percentile tuples
                to be considered.
            formulas (list of str): list of chemical formulas

        Returns:
            best_score, key, index (tuple): key is appropriate key in result.dict

        """
        max_score = [0, None, None, None]
        for key, i, entry in self.extract_results(
            molecules=molecules,
            charges=charges,
            file_names=file_names,
            label_percentiles=label_percentiles,
            formulas=formulas,
        ):
            if entry.score > max_score[0]:
                max_score[0] = entry.score
                max_score[1] = key
                max_score[2] = i
                max_score[3] = entry
        return max_score

    def determine_max_itensity(self, obj_for_calc_amount):
        """
        Function to determine the maximum intensity in given elution
        profile. The structure of the object passed to the function is shown
        below. This fucntion can be used as a template function to write and use
        of function to determine the amount of a molecule e.g. area under curve
        or summed up intensities. All self written function must return
        amount, rt at or around the amount and the mScore at or around the
        amount in an dictionary with appropiate key names.

        Example key names (default):

            * 'max I in window'
            * 'max I in window (rt)'
            * 'max I in window (score)'

        Note:

            This is the default function to determine the peptide amount when
            write_amount_csv() is called.

        Examples::

            {
               'rt'       : [rt1,rt2,...],
               'i'        : [in1,in2,...],
               'scores'   : [sc1,sc2,...],
               'spec_ids' : [id1,idt2,...],
            }

        Returns:

            dict: keys are shown above

        """
        return_dict = None
        if len(obj_for_calc_amount["i"]) != 0:
            maxI = max(obj_for_calc_amount["i"])
            index_of_maxI = obj_for_calc_amount["i"].index(maxI)
            amount_rt = obj_for_calc_amount["rt"][index_of_maxI]
            amount_score = obj_for_calc_amount["scores"][index_of_maxI]

            return_dict = {
                "max I in window": maxI,
                "max I in window (rt)": amount_rt,
                "max I in window (score)": amount_score,
            }
        return return_dict

    def _determine_measured_error(
        self,
        score_threshold=None,
        topX=3,
        filename=None,
        plot=True,
        formulas_to_test=None,
    ):
        """
        Function to determine and plot (density plot) the m/z an intensity
        error.

        """
        if filename == None:
            filename = "Histograms_mz_and_i_error.pdf"
        kwargs = {"score_threshold": score_threshold}
        if formulas_to_test is not None:
            kwargs["formulas"] = formulas_to_test

        error_dict = {
            "mz_error": [],
            "intensity_error": [],
            "time_dependent_mz_error": ddict(list),
            "time_dependent_intensity_error": ddict(list),
        }
        for key, i, entry in self.extract_results(**kwargs):
            rt = int(round(entry.rt))
            for mmz, mi, rel_i, cmz, ci in sorted(
                entry.peaks, key=itemgetter(2), reverse=True
            )[:topX]:
                if mmz is not None:
                    si = ci * entry.scaling_factor
                    rel_i_error = abs(mi - si) / si
                    if rel_i_error > 1:
                        rel_i_error = 1
                    rel_mz_error = (mmz - cmz) / cmz
                    rel_mz_error_in_ppm = rel_mz_error * 1e6

                    error_dict["mz_error"].append(rel_mz_error_in_ppm)
                    error_dict["intensity_error"].append(rel_i_error)
                    error_dict["time_dependent_mz_error"][rt].append(
                        rel_mz_error_in_ppm
                    )
                    error_dict["time_dependent_intensity_error"][rt].append(rel_i_error)
        if plot and len(error_dict["mz_error"]) > 0:
            assert self._import_rpy2() == True, "require R & rpy2 installed..."
            grdevices.pdf(filename)
            for plot_data_type, xaxis_label in [
                ("mz_error", "m/z error [ppm]"),
                ("intensity_error", "intensity error [rel.]"),
            ]:
                # for plot_data_type, value_object in error_dict.items():
                # for error_name, value_list in [('m/z error [ppm]', mz_error_list), ('intensity error [rel.]', i_error_list)]:
                value_object = error_dict[plot_data_type]
                if type(value_object) == type([]):

                    if len(value_object) > 1:
                        plot_data = r.density(
                            rpy2.robjects.vectors.FloatVector(value_object)
                        )
                        maxDensity = plot_data[0][plot_data[1].index(max(plot_data[1]))]
                        x_data = plot_data[0]
                        y_data = plot_data[1]
                        N = len(value_object)
                        mean = sum(value_object) / len(value_object)
                    else:
                        # plot_data = [ value_object[0], 1 ]
                        x_data = [value_object[0], value_object[0]]
                        y_data = [0, 1]
                        N = 1
                        mean = value_object[0]
                        maxDensity = mean

                    graphics.plot(
                        x_data,
                        y_data,
                        main="{0} N = {1}, mean = {2:1.3f} max = {3:1.3f}".format(
                            plot_data_type, N, mean, maxDensity
                        ),
                        xlab=xaxis_label,
                        ylab="density",
                        type="l",
                    )
                # plot histogram as well?
                # graphics.hist(
                # rpy2.robjects.vectors.FloatVector(i_error_list),
                #     main='rel i error',
                #     xlab='rel i error',
                #     ylab='density'
                # )
            grdevices.dev_off()
        return error_dict

    def _import_rpy2(self):
        """Imports all rpy2 related modules into global namespace."""
        try:
            global rpy2
            global importr
            global robjects
            global r
            global graphics
            global grdevices
            import rpy2
            from rpy2.robjects.packages import importr
            import rpy2.robjects as robjects
            from rpy2.robjects import r

            graphics = importr("graphics")
            grdevices = importr("grDevices")
            success = True
        except:
            success = False
        return success

    def _define_i_transformation(self, tag=None):
        if tag is None:
            tag = "default"
        functions = {
            "log10": lambda x: math.log(1 + x, 10),
            "log2": lambda x: math.log(1 + x, 2),
            "default": lambda x: x,
        }
        if tag not in functions.keys():
            print(
                "[ Warning ] Do not know tag {0} - falling back to default intensity transformation".format(
                    tag
                )
            )
            tag = "default"
        i_label = "Intensity [a.u.]"
        if tag is not "default":
            i_label = tag + " " + i_label
        return i_label, functions[tag]

    def _generate_r_colors(self, tag, elements):
        p = {
            "rainbow": list(r.rainbow(elements, start=0.2, end=1)),
            "terrain": list(r("{0}.colors".format("terrain"))(elements)),
            "topo": list(r("{0}.colors".format("cm"))(elements)),
            "heat": list(r("{0}.colors".format("heat"))(elements)),
        }
        reverse = False
        if tag[-2:] == "_r":
            reverse = True
            tag = tag[:-2]
        if tag not in p.keys():
            print("Do not {0} as color tag, falling back to rainbow")
            tag = "rainbow"
        colors = p[tag]
        if reverse:
            colors.reverse()
        return colors

    def _format_chemical_formula_for_r(self, key):
        formula = key.formula
        # print(formula)
        formated_formula = re.sub(
            r"(?P<element>[A-Z]{1}[a-z]*).(?P<count>[0-9]*).",
            lambda match: '"{0}" ["{1}"], '.format(
                match.group("element"), match.group("count")
            ),
            formula,
        )
        # print(formated_formula)
        formated_formula = re.sub(
            r'\((?P<isotop>[0-9]+)\)"(?P<element>[A-Z]{1}[a-z]*)"',
            lambda match: '""^"{0}","{1}" '.format(
                match.group("isotop"), match.group("element")
            ),
            formated_formula,
        )

        r_expression = r("expression(paste({0}))".format(formated_formula))
        return r_expression

    def _format_molecule_for_r(self, key):
        # molecule = '/'.join( self.lookup['formula to molecule'][key.formula] )
        molecule = self.lookup["formula to molecule"][key.formula][0]
        molecule = re.sub(r"[0-9]+", lambda x: '["' + x.group() + '"], ', molecule)
        molecule = re.sub(r"[a-zA-Z]+", lambda x: '"' + x.group() + '" ', molecule)
        molecule = '{0},""^"+{1}"'.format(molecule, key.charge)
        r_expression = r("expression(paste({0}))".format(molecule))
        return r_expression

    def plot_MIC_3D(self, key, file_name=None, rt_window=None, i_transform=None):
        """
        Plot MIC from results using rpy2 in 3D.



        """
        # rt_window = None
        assert self._import_rpy2() is True, "require R & rpy2 installed..."
        assert key in self.keys(), "key has no match results"

        i_label, i_trans_function = self._define_i_transformation(tag=i_transform)
        data = {}
        all_specIDs = set()
        all_mz_values = set()
        for entry in self[key]["data"]:
            if rt_window is None or rt_window[0] <= entry.rt <= rt_window[1]:
                all_specIDs.add(int(entry.spec_id))
                for mmz, mi, ri, cmz, ci in entry.peaks:
                    if mmz is None:
                        intensity = 0
                    else:
                        intensity = ci * entry.scaling_factor
                    data[int(entry.spec_id), cmz] = (intensity, entry.score)
                    all_mz_values.add(cmz)

        if len(all_mz_values) == 0:
            print("No matches found within window for key", key)
        else:
            min_mz = min(all_mz_values)
            max_mz = max(all_mz_values)
            for specID, mz in list(data.keys()):
                data[specID, min_mz - 0.1] = (0, data[specID, min_mz][1])
                data[specID, max_mz + 0.1] = (0, data[specID, max_mz][1])
                all_mz_values |= set([min_mz - 0.1, max_mz + 0.1])
                for var in [-0.2, +0.2]:
                    all_mz_values.add(mz + var)
                    data[specID, mz + var] = (0, data[specID, mz][1])
            x_values = sorted(all_specIDs)
            y_values = sorted(all_mz_values)
            z_values = []
            c_values = []
            if len(x_values) <= 3:
                print("Less than 3 x_values for key", key)
            else:
                colors = []
                for n in range(0, 101, 1):
                    rgb_col, hex_col = self.colorize_score(n / 100)
                    colors.append(hex_col)
                for y in y_values:  # mz
                    for x in x_values:  # specid
                        if (x, y) not in data.keys():
                            z = 0.0
                            c = colors[0]
                        else:
                            z = i_trans_function(data[(x, y)][0])
                            c = colors[int(round(data[(x, y)][1] * 100))]
                        if y != y_values[-1] and x != x_values[-1]:
                            c_values.append(c)
                        z_values.append(z)
                COLORS = {
                    "1_black": {"bg": "black", "fg": "white"},
                    "2_white": {"bg": "white", "fg": "black"},
                    "3_wwu": {"bg": "#EDEDED", "fg": "black"},
                }
                grdevices.pdf("{0}_MIC_3D.pdf".format(file_name))
                #
                # try:
                # # if True:
                #     formated_formula  = self._format_chemical_formula_for_r( key )
                #     formated_molecule = self._format_molecule_for_r( key )
                # except:
                if True:
                    # print("Warning - could not format {0} for r".format( key ))
                    formated_formula = "{0} charge: {1} lable percentile: {2}".format(
                        key.formula, key.charge, key.label_percentiles
                    )
                    formated_molecule = " ".join(
                        self.lookup["formula to molecule"][key.formula]
                    )
                for plottingType in sorted(COLORS.keys()):
                    # grdevices.png('test_MIC.png' , width = 1600, height = 1600)
                    bg = COLORS[plottingType]["bg"]
                    fg = COLORS[plottingType]["fg"]
                    graphics.par(
                        bg=bg,
                        fg=fg,
                        **{
                            "col.axis": fg,
                            "col.main": fg,
                            "col.lab": fg,
                            "col.sub": fg,
                            "cex.axis": 0.7,
                            "col.ticks": 5,
                            "las": 2,
                            # "lwd"       :   1,
                        }
                    )
                    graphics.persp(
                        robjects.FloatVector(x_values),
                        robjects.FloatVector(y_values),
                        r.matrix(
                            robjects.FloatVector(z_values),
                            nrow=len(x_values),
                            byrow=False,
                        ),
                        main=formated_molecule,
                        sub=formated_formula,
                        xlab="\n\nspectrum id",
                        ylab="m / z",
                        zlab="\n\n" + i_label,
                        ticktype="detailed",
                        nticks=5,
                        theta=+127,
                        phi=13,
                        # d      = 0.17,
                        r=1,
                        shade=0.95,
                        border="NA",  #'#00000021',
                        ltheta=-120,
                        lphi=40,
                        # ylim     = robjects.FloatVector([614,624]),
                        col=robjects.StrVector(c_values),
                    )
                    graphics.legend(
                        "topright",
                        title="mScore",
                        legend=r.c(
                            ["{0:2.1f}".format(i / 100.0) for i in range(0, 101, 10)]
                        ),
                        fill=robjects.StrVector([colors[i] for i in range(0, 101, 10)]),
                        bty="n",
                        xpd=True,
                        cex=0.7,
                    )
                grdevices.dev_off()
        return

    def colorize_score(self, score):
        """"""
        ##
        color = [0, 0, 0]  # copy becauuuuse  ?
        colorGradient = [
            (score_threshold, rgb_tuple)
            for (score_threshold, rgb_tuple) in sorted(self.params["COLORS"].items())
        ]
        if score is not None:
            idx = bisect.bisect(colorGradient, (score,))
            if idx == 0:
                color = colorGradient[0][1]
            elif idx == len(colorGradient):
                color = colorGradient[-1][1]
            else:
                # linear interpolation ... between idx-1 & idx
                dX = (score - colorGradient[idx - 1][0]) / (
                    colorGradient[idx][0] - colorGradient[idx - 1][0]
                )
                for color_chanel in range(3):
                    d_ = dX * (
                        colorGradient[idx][1][color_chanel]
                        - colorGradient[idx - 1][1][color_chanel]
                    )
                    if abs(d_) <= sys.float_info.epsilon:
                        color[color_chanel] = int(
                            round(colorGradient[idx - 1][1][color_chanel])
                        )
                    else:
                        color[color_chanel] = int(
                            round(colorGradient[idx - 1][1][color_chanel] + d_)
                        )
        hexed_color = "#" + "".join([hex(c)[2:] for c in color])

        return color, hexed_color

    def plot_MIC_2D(self, key, kwargs):
        """"""
        self.plot_MICs_2D([key], **kwargs)
        return

    def init_r_plot(self, file_name):
        assert self._import_rpy2() is True, "require R & rpy2 installed..."
        graphics = importr("graphics")
        grdevices = importr("grDevices")
        grdevices.pdf(file_name)
        return graphics, grdevices

    def plot_MICs_2D(
        self,
        key_list,
        file_name=None,
        rt_window=None,
        i_transform=None,
        xlimits=None,
        additional_legends=None,
        title=None,
        zlimits=None,
        ablines=None,
        graphics=None,
    ):
        """
        Args:
            additional_legends (dict): key points on lists of strings that are
                plotted as well.
        """

        assert self._import_rpy2() is True, "Require R plot .. use self.init_r_plot"
        if graphics is None:
            graphics, grdevices = self.init_r_plot(file_name)
        if zlimits is None:
            zlimits = [0, 1]
        zlimits_color_ints = [
            math.floor(zlimits[0] * 100.0),
            math.ceil(zlimits[-1] * 100.0),
        ]
        i_label, i_trans_function = self._define_i_transformation(tag=i_transform)
        graphics.par(
            mfrow=robjects.FloatVector([len(key_list), 1]),
            # mar = r.c( 2, 2, 1, 2 ),
            # oma = r.c( 2, 2, 2, 2 )
        )
        colors = []
        for n in range(0, 101, 1):
            rgb_col, hex_col = self.colorize_score(n / 100)
            colors.append(hex_col)
        # colors = self._generate_r_colors( 'rainbow', zlimits_color_ints[-1]-zlimits_color_ints[0]+1)
        for n, key in enumerate(key_list):
            if key not in self.keys():
                print("Warning, do not have match results for {0}".format(key))
                continue
            x = []
            y = []
            s = []
            c = []
            for entry in self[key]["data"]:
                if rt_window is None or rt_window[0] <= entry.rt <= rt_window[1]:
                    x.append(entry.rt)
                    y.append(entry.scaling_factor)
                    s.append(entry.score)
            assert (
                min(s) >= zlimits[0]
            ), "zlimits are set wrong, plots wont be conform, min score was {0}".format(
                min(s)
            )
            assert (
                max(s) <= zlimits[-1]
            ), "zlimits are set wrong, plots wont be conform, max score was {0}".format(
                max(s)
            )
            # min_score = math.floor(min(s) * 100)
            # max_score = math.ceil( max(s) * 100)
            for score in s:
                c.append(colors[int(round((score - zlimits[0]) * 100))])
            #     try:
            #     except:
            #         print(score, min_score, max_score, len(colors))
            max_y = max(y)
            # y = [ float(y_value) / float(max_y) for y_value in y ]
            if xlimits is None:
                if len(x) > 1:
                    xlimits = [x[0], x[-1]]
                else:
                    xlimits = [0, 120]
            if additional_legends is not None and key in additional_legends:
                additional_legend = additional_legends[key]
            else:
                additional_legend = None
            params = {
                # 'lwd' : 0.5,
                "pch": 19,
                "cex": 0.7,
                "xlab": "Retention Time [min]",
                "ylab": "Relative  Abundance",
                "xlim": r.c(xlimits[0], xlimits[-1]),
                # 'yaxt' : 'n',
                "ylim": r.c(0, max_y * 1.1),
                # 'ylas' : 2,
                "frame": False,
            }
            if n == 0 and title is not None:
                params["main"] = title
            if n != len(key_list) - 1:
                params["xaxt"] = "n"
            # print(params)
            graphics.plot(
                robjects.FloatVector(x),
                robjects.FloatVector(y),
                type="l",
                lwd=0.2,
                col="grey",
                **params
            )

            graphics.points(
                robjects.FloatVector(x),
                robjects.FloatVector(y),
                col=robjects.StrVector(c),
                lwd=0.1,
                **params
            )

            if xlimits[-1] - xlimits[0] < 10:
                step = 1
            else:
                step = 10
            for k in range(math.ceil(xlimits[0]), math.floor(xlimits[-1]), step):
                graphics.abline(
                    **{"v": float(k), "col": "gray", "lty": "dashed", "lwd": 0.4}
                )

            if additional_legend is not None:
                graphics = self._add_additional_legend_into_r_plot(
                    graphics, additional_legend
                )

            graphics = self._insert_mscore_legend_into_r_plot(
                graphics, colors, zlimits_color_ints
            )
            if ablines is not None:
                if key in ablines:
                    graphics = self._add_ablines_into_r_plot(graphics, ablines[key])

        return graphics

    def _add_ablines_into_r_plot(self, graphics, ablines):
        """
        ablines representing a list with dicts passed to the abline
        function of rpy2

        Example:
            ablines = [
                {
                    'v' : 1
                }
            ]

        The example above will plot a vertical line at x=1

        Please note that the default entries are:
            {
                'col' : 'black',
                'lty' : 'solid'
            }

        """
        defaultValues = {"col": "black", "lty": "solid"}
        for paramDict in ablines:
            for k, v in defaultValues.items():
                if k not in paramDict.keys():
                    paramDict[k] = v
            graphics.abline(**paramDict)
        return graphics

    def _add_additional_legend_into_r_plot(self, graphics, additional_legends):
        """
        Minimal a position with x and y is required as well as the label

        paramDict = {
            'x' :  0,
            'y' ;  0,
            'legend': 'origin'
        }

        """

        defaultValues = {"pos": 2, "cex": 0.5, "col": "black", "srt": 0}
        for paramDict in additional_legends:
            # defaultValues.update( paramDict )
            for k, v in defaultValues.items():
                if k not in paramDict.keys():
                    paramDict[k] = v
            graphics.text(
                paramDict["x"],
                paramDict["y"],
                labels=paramDict["text"],
                pos=paramDict["pos"],
                cex=paramDict["cex"],
                col=paramDict["col"],
                srt=paramDict["srt"]
                # adj=robjects.FloatVector([0,1]),
                # offset=robjects.FloatVector([0,-5]),
                # srt=-30,
            )

        return graphics

    def _insert_mscore_legend_into_r_plot(self, graphics, colors, zlimits_color_ints):
        graphics.legend(
            "topright",
            title="mScore",
            legend=r.c(
                [
                    "{0:2.1f}".format(i / 100.0)
                    for i in range(zlimits_color_ints[0], zlimits_color_ints[1] + 1, 10)
                ]
            ),
            fill=robjects.StrVector(
                [
                    colors[i]
                    for i in range(
                        0, zlimits_color_ints[-1] - zlimits_color_ints[0] + 1, 10
                    )
                ]
            ),
            bty="n",
            xpd=True,
            cex=0.7,
        )
        return graphics

    def determine_label_efficiency(self, element=None, key_translations=None):
        collector = {}
        t = {"percentile": "percentile", "mscore": "mscore", "count": "count"}
        if key_translations is not None:
            t.update(key_translations)
        for key, value_dict in self.items():
            for e, percentile in key.label_percentiles:
                if e == element:
                    break
            formated_percentile = round(float(percentile), 3)
            for entry in value_dict["data"]:
                formated_score = round(entry.score, 3)
                p_key = (formated_percentile, formated_score)
                try:
                    collector[p_key]
                except:
                    collector[p_key] = {
                        t["percentile"]: formated_percentile,
                        t["mscore"]: formated_score,
                        t["count"]: 0,
                    }
                collector[p_key][t["count"]] += entry.scaling_factor
        return collector.values()

    def _group_silac_pairs(self, silac_pairs=None):
        """
        Generate a list of silac formula pairs.

        Note:
        This will NOT recheck self.params['SILAC_AAS_LOCKED_IN_EXPERIMENT']
        all AAs in silac_pairs[ n ][0] will be flipped to [ n ][1]
        formulas that hold one [ n ][1] are not considered.
        """
        not_flipped = [target.upper() for source, target in silac_pairs]
        to_be_flipped = [source.upper() for source, target in silac_pairs]
        pairs = []
        # print(silac_pairs)
        for org_molecule, variants in self.lookup[
            "molecule fixed label variations"
        ].items():
            for variant in variants:
                has_target_aa = False
                for discared_AA_config in not_flipped:
                    if discared_AA_config in variant:
                        has_target_aa = True
                        break
                if has_target_aa:
                    continue
                has_source_aa = False
                for hot_AA_config in to_be_flipped:
                    if hot_AA_config in variant:
                        has_source_aa = True
                        break
                if not has_source_aa:
                    continue
                # print( variant )
                pairs.append([variant])
                for source, target in silac_pairs:
                    source_aa = source[0].upper()
                    source_state = source[1]
                    target_aa = target[0].upper()
                    target_state = target[1]
                    re_s = "(?P<SILAC>[{0}]{{1}})(?P<state>[{1}]*)".format(
                        source_aa, source_state
                    )
                    variant = re.sub(
                        r"" + re_s,
                        lambda match: "{0}{1}".format(target_aa, target_state),
                        variant,
                    )
                pairs[-1].append(variant)
        return pairs

    def _group_14N_15N_pairs(self):
        """Generate pairs of 14N and 15N m_keys"""
        for formula in self.index["formulas"]:
            for file_name in self.index["files"]:
                for charge in self.index["charges"]:
                    light_key = self._m_key_class(
                        file_name,
                        formula,
                        charge,
                        (("N", self.params["PERCENTILE_FORMAT_STRING"].format(0)),),
                    )
                    heavy_key = self._m_key_class(
                        file_name,
                        formula,
                        charge,
                        (("N", self.params["PERCENTILE_FORMAT_STRING"].format(0.99)),),
                    )
                    # print( light_key)
                    # print( heavy_key)
                    if light_key in self.keys() and heavy_key in self.keys():
                        yield light_key, heavy_key
                        # exit(1)

    def calc_amounts_from_rt_info_file(
        self,
        rt_info_file=None,
        rt_border_tolerance=None,
        calc_amount_function=None,
        evidence_score_field="PEP",
        buffer_only=False,
        buffered_csv_dicts=None,
    ):
        """
        Function to calculate molecule/peptide amounts based on the quant
        summary/rt info file genearte by :py:func:`write_rt_info_file`.
        See e.g. example script generate_quant_summary_file.py. A function to
        calculate the final molecule amounts can be defined otherwise the
        default maximum intensity function is used.

        Args:
            rt_info_file (str): output file name of the quant summary/rt info
                csv file, must be be a complete or relative path
            rt_border_tolerance (int): retention time border tolerance in
                minutes
            calc_amount_function (obj): python function to calculate final
                amounts based on a simple dictionary structure


        The function to calculate the amount of the molecules
        (calc_amount_function) should be able to process the below shown
        dictionary structure (obj_for_calc_amount). The default function returns
        the maximum amount in the retention time window or in the complete
        profile. The function should return the determined amount, the retention
        time (or approximate) as well as the score.
        If functions are used which deternine the amount over more than one
        spectrum retention times for this amount should be e.g. at the maximum
        intensity of the profile. Scores could be e.g. averaged or also the
        score at the maximum amount could be used.

        Example for obj_for_calc_amount::

            {
                'rt'       : [rt1,rt2,...],
                'i'        : [in1,in2,...],
                'scores'   : [sc1,sc2,...],
                'spec_ids' : [id1,idt2,...],
            }

        Example key names (default):

            * 'max I in window'
            * 'max I in window (rt)'
            * 'max I in window (score)'
            * 'auc in window' (area under curve)
            * 'sum I in window' (summed up intensities)


        """
        if rt_info_file is not None:

            tmp_csv_dicts = []
            # csv_reader
            if rt_info_file.endswith(".csv"):
                with codecs.open(rt_info_file, mode="r", encoding="utf-8") as rif:
                    dict_reader = csv.DictReader(rif)
                    for line_dict in dict_reader:
                        tmp_csv_dicts.append(line_dict)
            elif rt_info_file.endswith(".xlsx"):
                # read xlsx
                tmp_csv_dicts = pyqms.adaptors.read_xlsx_file(rt_info_file)
            else:
                print(
                    "Extension: {0} of file {1} not recognized".format(
                        rt_info_file.split(".")[-1], rt_info_file
                    )
                )
                exit(1)
        elif buffered_csv_dicts is not None:
            tmp_csv_dicts = buffered_csv_dicts
        else:
            print("No RT info file or buffered csv lines were provided. Exiting now")
            exit()
        list_of_csvdicts = []
        for line_dict in tmp_csv_dicts:
            haz_rt_info = True
            for imp_key in ["start (min)", "stop (min)"]:
                imp_field = line_dict.get(imp_key, None)
                try:
                    line_dict[imp_key] = float(imp_field)
                except:
                    haz_rt_info = False
            if haz_rt_info:
                # print(line_dict)
                # exit()
                if type(line_dict["label_percentiles"]) is str:
                    label_percentiles = eval(line_dict["label_percentiles"])
                else:
                    label_percentiles = line_dict["label_percentiles"]

                m_key = self._m_key_class(
                    line_dict["file_name"],
                    line_dict["formula"],
                    int(line_dict["charge"]),
                    label_percentiles,
                )
                obj_for_calc_amount = {"rt": [], "i": [], "scores": [], "spec_ids": []}
                # calculate the lists and pass to the calc amoutn fucntion...
                for entry in self[m_key]["data"]:
                    try:
                        if entry.rt[1] == "second":
                            rt = entry.rt[0] / 60
                        elif entry.rt[1] == "minute":
                            rt = entry.rt[0]
                    except TypeError:
                        rt = entry.rt

                    if rt < line_dict["start (min)"]:  # i.e. start of rt window
                        continue
                    elif rt > line_dict["stop (min)"]:  # i.e. end of rt window
                        break
                    else:
                        obj_for_calc_amount["rt"].append(rt)
                        obj_for_calc_amount["i"].append(entry.scaling_factor)
                        obj_for_calc_amount["scores"].append(entry.score)
                        obj_for_calc_amount["spec_ids"].append(entry.spec_id)

                # if len(obj_for_calc_amount['i']) < min_profile_length:
                #     #check that at least one spec was added
                #     continue
                if calc_amount_function is None:
                    amount_dict = self.determine_max_itensity(obj_for_calc_amount)
                else:
                    amount_dict = calc_amount_function(obj_for_calc_amount)
                # print(amount_dict)
                # exit()
                if amount_dict is not None:
                    line_dict.update(amount_dict)
            list_of_csvdicts.append(line_dict)
        if buffer_only is False:
            self.write_rt_info_file(
                output_file=rt_info_file,
                list_of_csvdicts=list_of_csvdicts,
                rt_border_tolerance=rt_border_tolerance,
                update=False,
                buffer_only=buffer_only,  # default False
            )
        return list_of_csvdicts

    def _determine_rt_windows_from_evidence(self, rt_border_tolerance=None):
        if rt_border_tolerance is None:
            rt_border_tolerance = 0
        rt_border_lookup = {}
        molecule_lookup = {}
        if "formula to evidences" in self.lookup.keys():
            for formula in self.lookup["formula to evidences"].keys():
                if len(self.lookup["formula to evidences"]) != 0:
                    tmp_evidence_dict = self.lookup["formula to evidences"].get(
                        formula, None
                    )
                    if tmp_evidence_dict is not None:
                        molecule_lookup[formula] = tmp_evidence_dict.keys()
                        # curate windows here
                        rt_border_lookup[formula] = self.curate_rt_windows(
                            tmp_evidence_dict, rt_tolerance=rt_border_tolerance
                        )
        return rt_border_lookup, molecule_lookup

    def write_rt_info_file(
        self,
        output_file=None,
        list_of_csvdicts=None,
        trivial_name_lookup=None,
        rt_border_tolerance=None,
        update=True,
        buffer_only=False,
    ):
        """
        Function to write a default quant summary/rt info file. See e.g.
        example script generate_quant_summary_file.py.

        Args:
            output_file (str): output file name of the csv, should be a
                complete path
            list_of_csvdicts (list): list of dictionaries passed to the
                DictWriter class, default fieldnames can be found below
            trivial_name_lookup (dict): self defined trivial_name_lookup, see
                format below.
            rt_border_tolerance (int): retention time border tolerance in
                minutes
            update (bool): if True read in or passed dictionaries in
                list_of_csvdicts will be updated with default evidence and
                trivial name information

        The quant summary file can manually be updated (e.g. the start and stop
        RT information). If an evidence lookup is present in the result class (
        can be passed to the isotopologue library or later be set in the result
        class), these information are used to define the retention time borders
        (e.g. peptide identfication information from peptide spectrum matches).

        Default fieldnames:

            * file_name               : filename of spectrum input file
            * formula                 : molecular formula of the molecule
            * molecule                : molecule or trivial name
            * trivial_name(s)         : protein or trivial names
            * label_percentiles       : labeling percentile ( (element, enrichment in %), )
            * charge                  : charge of the molecule
            * start (min)             : start of retention time window
            * stop (min)              : stop of retention time window
            * max I in window         : maximum intensity in retention time window
            * max I in window (rt)    : retention time @ maximum intensity in retention time window
            * max I in window (score) : score @ maximum intensity in retention time window
            * auc in window           : area under curve in retention time window
            * sum I in window         : summed up intensities in retention time window
            * evidences (min)         : all evidences/identifications (score@rt;...)

        Trivial name lookup example::

            {
                'C(33)H(59)14N(1)N(8)O(9)S(1)' : ['BSA','Bovine serum albumine']
            }

        """
        assert output_file is not None, "You need to specify an output file"

        if rt_border_tolerance is None:
            rt_border_tolerance = 0

        if list_of_csvdicts is None:
            list_of_csvdicts = []
            for m_key in sorted(self.keys()):
                tmp = m_key._asdict()
                list_of_csvdicts.append(tmp)

        if trivial_name_lookup is None:
            trivial_name_lookup = {}

        csv_kwargs = {"extrasaction": "ignore"}
        if sys.platform == "win32":
            csv_kwargs["lineterminator"] = "\n"
        else:
            csv_kwargs["lineterminator"] = "\r\n"

        (
            full_rt_border_lookup,
            full_molecule_lookup,
        ) = self._determine_rt_windows_from_evidence(
            rt_border_tolerance=rt_border_tolerance
        )
        lines_2_write = []
        for tmp in list_of_csvdicts:
            if update:
                trivial_names = trivial_name_lookup.get(tmp["formula"], [])
                current_trivial_name = tmp.get("trivial_name(s)", None)
                if current_trivial_name is None:
                    tmp["trivial_name(s)"] = ", ".join(trivial_names)
                elif current_trivial_name != "":
                    tmp["trivial_name(s)"] += ", ".join(trivial_names)
                else:
                    pass

                evidence_lookup_present = False
                if tmp["formula"] in full_rt_border_lookup.keys():
                    if len(full_rt_border_lookup[tmp["formula"]].keys()) > 0:
                        rt_border_lookup = full_rt_border_lookup[tmp["formula"]]
                        molecule_list = full_molecule_lookup[tmp["formula"]]
                        evidence_lookup_present = True
                        tmp_evidence_dict = self.lookup["formula to evidences"].get(
                            tmp["formula"], None
                        )

                if evidence_lookup_present is False:
                    molecule_list = self.lookup["formula to molecule"].get(
                        tmp["formula"], []
                    )
                for molecule in molecule_list:
                    tmp["molecule"] = molecule
                    if evidence_lookup_present is True:
                        # 'start (min)',
                        # 'stop (min)',
                        # 'evidences (min)'
                        # all_evidence_rts = []
                        evidence_info_string_list = []
                        for evidence_info_dict in tmp_evidence_dict[molecule][
                            "evidences"
                        ]:
                            """
                            {
                                'RT'          : float(line_dict['Retention Time (s)']) / 60.0, # always in min
                                'score'       : float(line_dict['PEP']),
                                'score_field' : 'PEP',

                            }
                            """
                            if "RT" not in evidence_info_dict.keys():
                                continue
                            # all_evidence_rts.append( evidence_info_dict['RT'] )

                            if evidence_info_dict["score"] != "None":
                                evidence_string = "{0}@{1}".format(
                                    evidence_info_dict["score"],
                                    evidence_info_dict["RT"],
                                )
                            else:
                                evidence_string = "{0}".format(evidence_info_dict["RT"])

                            evidence_info_string_list.append(
                                (evidence_info_dict["RT"], evidence_string)
                            )

                        # get info from rt_border_lookup, do it like this:
                        if len(rt_border_lookup.keys()) > 0:
                            lower_border_tolerance = rt_border_lookup[molecule].get(
                                "lower_window_border", rt_border_tolerance
                            )
                            upper_border_tolerance = rt_border_lookup[molecule].get(
                                "upper_window_border", rt_border_tolerance
                            )
                            tmp["start (min)"] = (
                                rt_border_lookup[molecule]["rt_window"][0]
                                - lower_border_tolerance
                            )
                            tmp["stop (min)"] = (
                                rt_border_lookup[molecule]["rt_window"][1]
                                + upper_border_tolerance
                            )

                        # old versio of rt window definition
                        # if len( all_evidence_rts) != 0:
                        #     tmp['start (min)'] = min(all_evidence_rts) - rt_border_tolerance
                        #     tmp['stop (min)'] = max(all_evidence_rts) + rt_border_tolerance

                        tmp["evidences (min)"] = ";".join(
                            [i[1] for i in sorted(evidence_info_string_list)]
                        )
                        if len(tmp_evidence_dict[molecule]["trivial_names"]) > 0:
                            tmp["trivial_name(s)"] = ";".join(
                                sorted(
                                    list(
                                        set(
                                            tmp_evidence_dict[molecule]["trivial_names"]
                                        )
                                    )
                                )
                            )
                        # if len()
                    lines_2_write.append(copy.deepcopy(tmp))
                    # csv_output.writerow( tmp )
            else:
                lines_2_write.append(copy.deepcopy(tmp))
                # csv_output.writerow( tmp )
        # print(lines_2_write)
        # default, write csv
        if buffer_only is False:
            if output_file.endswith(".csv"):
                with codecs.open(output_file, mode="w", encoding="utf-8") as infof:
                    csv_output = csv.DictWriter(
                        infof, default_amount_csv_fieldnames, **csv_kwargs
                    )
                    csv_output.writeheader()
                    for tmp in lines_2_write:
                        csv_output.writerow(tmp)
            elif output_file.endswith("xlsx"):
                # write xlsx
                try:
                    from openpyxl import Workbook
                except:
                    print("openpyxl is not installed, please install it and try again")
                    print("pip3.4 install openpyxl")
                    sys.exit(1)
                wb = Workbook()
                ws = wb.active

                for header_index, header in enumerate(default_amount_csv_fieldnames):
                    header_cell = ws.cell(row=1, column=header_index + 1)
                    header_cell.value = header

                row_counter = 2
                for tmp in lines_2_write:
                    for column, key in enumerate(default_amount_csv_fieldnames):
                        cell = ws.cell(row=row_counter, column=column + 1)
                        value_2_write = tmp.get(key, "")
                        cell.value = str(value_2_write)
                    row_counter += 1

                wb.save(output_file)
            else:
                print(
                    "Extension: {0} of file {1} not recognized".format(
                        output_file.split(".")[-1], output_file
                    )
                )
        return lines_2_write

    def write_result_csv(self, output_file_name=None):
        """
        Write raw results into a .csv file

        Args:
            output_file_name (str): output file name of the csv containing
                containing all raw results, should be a complete path

        Warning:

            Depending on data size the resulting csv can become very large.
            Some csv viewer can not handle files with a large number of
            lines.

        Keys in csv:

            * Formula           : molecular formula of the molecule (str)
            * Molecule          : molecule or trivial name (str)
            * Charge            : charge of the molecule (int)
            * ScanID            : ScanID of the quantified spectrum (int)
            * Label Percentiles : Labeling percentile ( (element, enrichment in %), )
            * Amount            : the determined amount of the molecule
            * Retention Time    : retetention time of the ScanID
            * mScore            : score of the isotopologue match
            * Filename          : filename of spectrum input files

        """
        raw_amounts_fieldnames = [
            "formula",
            "molecule",
            "charge",
            "scan_id",
            "label_percentiles",
            "intensity",
            "retention_time",
            "mScore",
            "file_name",
            "trivial_name(s)",
            "#exp. peaks",
            "#obs. peaks",
        ]
        map_formulas = False
        if len(self.lookup["formula to molecule"].values()) > 1:
            map_formulas = True

        evidence_lookup_present = False
        if "formula to evidences" in self.lookup.keys():
            evidence_lookup_present = True
        if output_file_name is None:
            output_file_name = "pyQms_results.csv"
        with codecs.open(output_file_name, mode="w", encoding="utf-8") as out_csv:
            csv_out = csv.DictWriter(out_csv, raw_amounts_fieldnames)
            csv_out.writeheader()
            for key, v_list in self.items():
                if evidence_lookup_present:
                    tmp_evidence_dict = self.lookup["formula to evidences"].get(
                        key.formula, None
                    )
                else:
                    tmp_evidence_dict = None
                for v in v_list["data"]:
                    dict2write = {
                        "formula": key.formula,
                        "molecule": None,
                        "charge": key.charge,
                        "scan_id": v.spec_id,
                        "label_percentiles": key.label_percentiles,
                        "intensity": v.scaling_factor,
                        "retention_time": v.rt,
                        "mScore": v.score,
                        "file_name": key.file_name,
                        "#exp. peaks": len(v.peaks),
                        "#obs. peaks": len([x for x in v.peaks if x[0] is not None]),
                    }
                    if map_formulas is False:
                        csv_out.writerow(dict2write)
                    else:
                        for molecule in self.lookup["formula to molecule"][key.formula]:
                            dict2write["molecule"] = molecule
                            if tmp_evidence_dict is not None:
                                if molecule in tmp_evidence_dict.keys():
                                    dict2write["trivial_name(s)"] = ";".join(
                                        tmp_evidence_dict[molecule]["trivial_names"]
                                    )
                            else:
                                trivial_names = self.lookup[
                                    "formula to trivial name"
                                ].get(key.formula, None)
                                if trivial_names is not None:
                                    dict2write["trivial_name(s)"] = ";".join(
                                        trivial_names
                                    )

                            csv_out.writerow(dict2write)
        return

    def write_result_mztab(self, output_file_name=None, rt_border_tolerance=None):
        """
            Write minimal peptide quantification results into a .mztab file. It
            is neccessary to specify the 'formula to evidences' dict in the
            lookup of the results class to write results!

            Note:

                This basic mzTab writer is still in beta stage. Use and
                evaluate with care.



            PRIDE CV based quantifcation unit and value is fixd to:

                * PRIDE:0000393, Relative quantification unit
                * PRIDE:0000425, MS1 intensity based label-free quantification method

            Args:
                output_file_name (str): output file name of the mztab containing
                    containing all raw results, should be a complete path

            Note:

                Adiitional information has to be passed tot he result class for
                a more complete mztab output.

            Keys in mztab:

                * sequence
                * accession
                * unique
                * database
                * database_version
                * search_engine
                * best_search_engine_score[1-n]
                * modifications
                * retention_time
                * retention_time_window
                * charge
                * mass_to_charge
                * peptide_abundance_study_variable[1-n]
                * peptide_abundance_stdev_study_variable[1-n]
                * peptide_abundance_std_error_study_variable[1-n]
                * search_engine_score[1-n]_ms_run[1-n]
                * peptide_abundance_assay[1-n]
                * spectra_ref
                * opt_{identifier}_*
                * reliability
                * uri

        Addtional information can be added to the mzTab file by adding a dict
        like shown below to the results.lookup dict under the key
        'mztab_meta_info'.: ::

            mztab_meta_info = {
                'protein_search_engine_score'   : [],
                'psm_search_engine_score'       : ['[MS,MS:1001475,OMSSA:evalue, ]'],
                'fixed_mod'                     : ['[UNIMOD, UNIMOD:4, Carbamidomethyl, ]'],
                'variable_mod'                  : ['[UNIMOD, UNIMOD:35, Oxidation, ]'],
                'study_variable-description'    : ['Standard BSA measurement'],
                'ms_run-location'               : ['BSA1.mzML'],
            }


        """
        assert "formula to evidences" in self.lookup.keys(), (
            'Please specify the "formula to evidences" dict in the lookup of '
            "the results class"
        )
        if rt_border_tolerance is None:
            rt_border_tolerance = 0
        default_value = "null"
        """
        COM Report for quantification results of pyQms
        MTD mzTab-version   1.0.0
        MTD mzTab-mode  Summary
        MTD mzTab-type  Quantification
        MTD description mzTab quantification results on peptide level
        MTD protein_search_engine_score[1]  [MS,MS:1001171,Mascot:score,]
        MTD psm_search_engine_score[1]  [MS,MS:1001171,Mascot:score,]
        MTD ms_run[1]-location  file://C:/path/to/my/file1.mzML
        MTD ms_run[2]-location  file://C:/path/to/my/file2.mzML
        MTD ms_run[3]-location  file://C:/path/to/my/file3.mzML
        MTD ms_run[4]-location  file://C:/path/to/my/file4.mzML
        MTD ms_run[5]-location  file://C:/path/to/my/file5.mzML
        MTD ms_run[6]-location  file://C:/path/to/my/file6.mzML
        MTD protein-quantification_unit [PRIDE, PRIDE:0000393, Relative quantification unit,]
        MTD fixed_mod[1]    [UNIMOD, UNIMOD:4, Carbamidomethyl, ]
        MTD variable_mod[1] [UNIMOD, UNIMOD:35, Oxidation, ]
        MTD study_variable[1]-description   heat shock response of control
        MTD study_variable[2]-description   heat shock response of treatment

        """
        HEADER = """COM Report for quantification results of pyQms
MTD mzTab-version   1.0.0
MTD mzTab-mode  Summary
MTD mzTab-type  Quantification
MTD description mzTab based pyQms quantification results on peptide level
MTD peptide-quantification-unit [PRIDE, PRIDE:0000393, Relative quantification unit,]
MTD peptide-quantification-value [PRIDE, PRIDE:0000425, MS1 intensity based label-free quantification method,]"""

        if "mztab_meta_info" not in self.lookup.keys():
            self.lookup["mztab_meta_info"] = {}
            if "ms_run-location" not in self.lookup["mztab_meta_info"].keys():
                self.lookup["mztab_meta_info"]["ms_run-location"] = []
                for file_name in self.index["files"]:
                    self.lookup["mztab_meta_info"]["ms_run-location"].append(file_name)

        header_info = {"assay": []}
        if "mztab_meta_info" in self.lookup.keys():
            for meta_info_key, meta_info_value_list in self.lookup[
                "mztab_meta_info"
            ].items():
                if "-" in meta_info_key:
                    meta_info_key, add_on_definition = meta_info_key.split("-")
                    add_on_definition = "-{0}".format(add_on_definition)
                    header_key = "{0}{1}".format(meta_info_key, add_on_definition)
                else:
                    add_on_definition = ""
                    header_key = meta_info_key
                if header_key not in header_info.keys():
                    header_info[header_key] = []
                for pos, meta_info_value in enumerate(meta_info_value_list):
                    header_info[header_key].append(
                        "MTD\t{0}[{1}]{2}\t{3}".format(
                            meta_info_key, pos + 1, add_on_definition, meta_info_value
                        )
                    )
            # add runs manually
        assay_ref_lookup = {}
        for pos, ms_run_ref in enumerate(header_info["ms_run-location"]):
            line_start, definition, value = ms_run_ref.split("\t")
            # pos, defintion = pos_and_definition.split('-')
            assay_identifier = "assay[{0}]".format(pos + 1)
            header_info["assay"].append(
                "MTD\t{0}-ms_run_ref\t{1}".format(assay_identifier, definition)
            )
            assay_ref_lookup[value] = assay_identifier

        # use already built in functions to determine windows, amounts etc.
        tmp_csv_dicts = self.write_rt_info_file(
            output_file="override",
            list_of_csvdicts=None,
            trivial_name_lookup=None,
            rt_border_tolerance=rt_border_tolerance,
            update=True,
            buffer_only=True,
        )

        tmp_csv_dicts = self.calc_amounts_from_rt_info_file(
            rt_info_file=None,
            rt_border_tolerance=None,
            calc_amount_function=None,
            # evidence_score_field = 'PEP',
            buffer_only=True,
            buffered_csv_dicts=tmp_csv_dicts,
        )

        mztab_fieldnames = [
            "PEH",
            "sequence",
            "accession",
            # 'unique',
            # 'database',
            # 'database_version',
            # 'search_engine',
            # 'best_search_engine_score[1-n]',
            "modifications",
            "retention_time",
            "retention_time_window",
            "charge",
            "mass_to_charge",
            # 'peptide_abundance_study_variable[1-n]',
            # 'peptide_abundance_stdev_study_variable[1-n]',
            # 'peptide_abundance_std_error_study_variable[1-n]',
            # 'search_engine_score[1-n]_ms_run[1-n]',
            # 'peptide_abundance_assay[1-n]',
            # 'spectra_ref',
            # 'opt_{identifier}_*',
            # 'reliability',
            # 'uri',
        ]

        for file_name, assay_identifier in assay_ref_lookup.items():
            mztab_fieldnames.append("peptide_abundance_{0}".format(assay_identifier))

        # pprint.pprint(tmp_csv_dicts)
        # exit()
        with open(output_file_name, "w") as io:
            print(HEADER, file=io)
            for header_key, header_add_on in header_info.items():
                if len(header_add_on) > 0:
                    print(os.linesep.join(header_add_on), file=io)
            print("", file=io)
            tsv_writer = csv.DictWriter(io, delimiter="\t", fieldnames=mztab_fieldnames)
            tsv_writer.writeheader()
            for line_dict in tmp_csv_dicts:
                if "max I in window" not in line_dict.keys():
                    continue
                dict2write = {fieldname: "null" for fieldname in mztab_fieldnames}
                if "#" in line_dict["molecule"]:
                    sequence, modification = line_dict["molecule"].split("#")
                else:
                    sequence = line_dict["molecule"]
                    modification = "null"
                dict2write = {
                    "PEH": "PEP",
                    "sequence": sequence,
                    "modifications": modification,
                    "retention_time_window": "{0:1.2f}|{1:1.2f}".format(
                        line_dict["start (min)"] * 60,  # in seconds
                        line_dict["stop (min)"] * 60,  # in seconds
                    ),
                    "charge": line_dict["charge"],
                    "retention_time": line_dict["max I in window (rt)"],
                    # those are added below
                    # 'peptide_abundance_assay[1-n]',
                    # those could be added with more info regarding the peptides
                    # 'accession',
                    # 'unique',
                    # 'database',
                    # 'database_version',
                    # 'search_engine',
                    # 'best_search_engine_score[1-n]',
                    # 'search_engine_score[1-n]_ms_run[1-n]',
                    # those could be added from the lib info
                    # 'mass_to_charge' :,
                    # these below we don not have/map at the moment :(
                    # 'peptide_abundance_study_variable[1-n]',
                    # 'peptide_abundance_stdev_study_variable[1-n]',
                    # 'peptide_abundance_std_error_study_variable[1-n]'
                    # 'spectra_ref',
                    # 'opt_{identifier}_*',
                    # 'reliability',
                    # 'uri',
                }
                abundance_key = "peptide_abundance_{0}".format(
                    assay_ref_lookup[line_dict["file_name"]]
                )
                dict2write[abundance_key] = line_dict["max I in window"]

                tsv_writer.writerow(dict2write)
                """
                {
                    'file_name': 'BSA1.mzML',
                    'formula': 'C(58)H(96)N(16)O(18)',
                    'charge': 2,
                    'label_percentiles': (('N', '0.000'),),
                    'trivial_name(s)': 'P02769',
                    'molecule': 'HLVDEPQNLIK',
                    'start (min)': 38.265035,
                    'stop (min)': 41.505281583333336,
                    'evidences (min)': '38.265035;38.30481363333333;40.39071451666667;41.46730143333333;41.505281583333336',
                    'max I in window (rt)': 41.492350260416664,
                    'max I in window': 26.547503653993502,
                    'max I in window (score)': 0.9593249834808004
                }
                """
        return

    def _smooth_list(self, input_list, k=5):
        """
        Smooth list with sliding window of k points
        """
        k = k
        fill_value = 0
        if k % 2 == 0:
            window_extention = 1
        else:
            window_extention = 0
        smoothend_list = [0 for entry in input_list]
        for pos in range(len(input_list)):
            local_smoothend_list = []
            lower_offset_range = -1 * math.floor(k / 2.0)
            upper_offset_range = +1 * math.ceil(k / 2.0)
            for offset in range(
                lower_offset_range, upper_offset_range + window_extention
            ):
                # print( offset )
                if pos + offset < 0 or pos + offset >= len(input_list):
                    continue
                value = input_list[pos + offset]
                local_smoothend_list.append(value)
            if len(local_smoothend_list) == 0:
                local_smoothend_list.append(0)
            # print(local_smoothend_list)
            # average = None
            if len(local_smoothend_list) != k:
                difference = k - len(local_smoothend_list)
                # average = sum(local_smoothend_list)/ float(len(local_smoothend_list))
                for missing_value in range(0, difference):
                    local_smoothend_list.append(0)  #  average )
            # print(pos, local_smoothend_list, average, k)
            smoothend_list[pos] = sum(local_smoothend_list) / float(k)
        # data_array                 = self[ key ]['data']
        # rts                        = [ entry.rt for entry in data_array ]
        # scores                     = [ entry.score for entry in data_array ]
        # intensities                = [ entry.scaling_factor for entry in data_array ]
        # max_score                  = max( scores )
        # max_score_index            = scores.index( max_score )
        # rt_at_max_score            = rts[ max_score_index ]
        # index_at_lower_border      = bisect.bisect( rts, rt_at_max_score - (float(window_length) / 2.))
        # if index_at_lower_border < 0:
        #     index_at_lower_border = 0
        # index_at_upper_border = bisect.bisect( rts, rt_at_max_score + (float(window_length) / 2.))
        # if index_at_upper_border > len( data_array ):
        #     index_at_upper_border = len(data_array)

        # exit(1)
        return smoothend_list

    def curate_rt_windows(self, evidence_dict, rt_tolerance):
        """
        Internal function to curate RT windows

        """
        # pprint.pprint(tmp_evidence_dict)
        window_sort_list = []
        for upep in evidence_dict.keys():
            rt_list = []
            for ident_dict in evidence_dict[upep]["evidences"]:
                rt_list.append(ident_dict["RT"])
            # q1 = numpy.percentile( rt_list, 25 )
            # q3 = numpy.percentile( rt_list, 75 )
            window_sort_list.append(
                (
                    [min(rt_list), max(rt_list)],
                    upep,
                    # [q1, q3]
                )
            )
        rt_border_lookup = {}
        if len(window_sort_list) > 1:
            # only fo this if we have two upeps for the same formula!
            rt_border_lookup = self.window_curator(
                window_sort_list, rt_tolerance=rt_tolerance
            )
        else:
            rt_border_lookup[window_sort_list[0][1]] = {
                "rt_window": window_sort_list[0][0]
            }
        return rt_border_lookup

    def window_curator(self, window_sort_list, rt_tolerance=None, win_key=None):
        window_sort_list.sort()
        upep_2_rt = {}
        for pos, current_window_object in enumerate(window_sort_list):
            # print(pos, current_window_object)
            # set windows for everz peptide!
            upep_2_rt[current_window_object[1]] = {
                "rt_window": current_window_object[0]
            }
            if pos == 0:
                continue
            last_window_object = window_sort_list[pos - 1]

            current_upep = current_window_object[1]
            current_window = current_window_object[0]
            # current_new_win    = current_window_object[2]

            last_upep = last_window_object[1]
            last_window = last_window_object[0]
            # last_new_win       = last_window_object[2]

            if current_window[0] - rt_tolerance <= last_window[-1] + rt_tolerance:

                # define new windows?
                if current_window[0] - rt_tolerance <= last_window[-1] + rt_tolerance:

                    max_diff_between_redefined_windows = (
                        current_window[0] - last_window[-1]
                    )
                    half_diff_redef_win = max_diff_between_redefined_windows / 2.0

                    current_upep_lower_border = upep_2_rt[current_upep].get(
                        "lower_window_border", rt_tolerance
                    )
                    current_upep_upper_border = upep_2_rt[current_upep].get(
                        "upper_window_border", rt_tolerance
                    )

                    last_upep_lower_border = upep_2_rt[last_upep].get(
                        "lower_window_border", rt_tolerance
                    )
                    last_upep_upper_border = upep_2_rt[last_upep].get(
                        "upper_window_border", rt_tolerance
                    )

                    unseparable = False
                    if (
                        current_window[0] - half_diff_redef_win
                        > current_window[1] + current_upep_upper_border
                    ):
                        unseparable = True
                    elif (
                        last_window[-1] - half_diff_redef_win
                        < last_window[0] - last_upep_lower_border
                    ):
                        unseparable = True
                    elif (
                        last_window[0] - last_upep_lower_border
                        < current_window[0] - current_upep_lower_border
                    ):
                        if (
                            last_window[-1] + last_upep_upper_border
                            > current_window[-1] + current_upep_upper_border
                        ):
                            # current windows encloses last winodw
                            unseparable = True

                    else:
                        pass

                    if unseparable is True:
                        if (
                            "window_is_unseparable"
                            not in upep_2_rt[current_upep].keys()
                        ):
                            upep_2_rt[current_upep]["window_is_unseparable"] = []
                        if "window_is_unseparable" not in upep_2_rt[last_upep].keys():
                            upep_2_rt[last_upep]["window_is_unseparable"] = []

                        upep_2_rt[current_upep]["window_is_unseparable"].append(win_key)
                        upep_2_rt[last_upep]["window_is_unseparable"].append(win_key)

                        for del_key in ["lower_window_border", "upper_window_border"]:
                            for p in [last_upep, current_upep]:
                                if del_key in upep_2_rt[p].keys():
                                    del upep_2_rt[p][del_key]
                    else:
                        upep_2_rt[current_upep][
                            "lower_window_border"
                        ] = half_diff_redef_win

                        upep_2_rt[last_upep][
                            "upper_window_border"
                        ] = half_diff_redef_win

        return upep_2_rt


if __name__ == "__main__":
    print(__doc__)
